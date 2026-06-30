#!/usr/bin/env python3
"""Summarize Cortellis gene XLSX into count tables for report "Summary" sections.

Inputs (preferred contract):
  - results/cortellis_<indication>/gene_cortellis_data.xlsx
  - sheet: Drugs_Comprehensive
  - columns to count:
      - Highest Phase
      - Primary Indications  (may be multi-valued; split on ';' or '|')

This script is a helper for report generators to build `gene.overview_html`:
it extracts:
  1) `Highest Phase` counts
  2) `Primary Indications` counts
  3) a contingency table (top indications × highest phase) that a report generator
     can render as a **stacked bar plot** (each bar = indication; stacks = phase).
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
import re
from typing import Iterable, List, Optional, Sequence, Tuple, Dict


def _find_gene_column(cols: Sequence[str]) -> Optional[str]:
    candidates = [
        "Gene",
        "GENE",
        "Gene Symbol",
        "Gene_Symbol",
        "Target",
        "Target Gene",
        "Target_Gene",
        "Target Symbol",
        "Target_Symbol",
    ]
    colset = {c.lower(): c for c in cols}
    for want in candidates:
        got = colset.get(want.lower())
        if got:
            return got
    return None


def _split_multi(v: str) -> List[str]:
    if not v:
        return []
    # Default contract: semicolon-delimited in Cortellis exports.
    # Avoid splitting on commas because disease names may contain commas.
    parts = re.split(r"[;|]", str(v))
    out = []
    for p in parts:
        p2 = p.strip()
        if p2.lower() == "nan":
            continue
        if p2:
            out.append(p2)
    return out


def _to_md_table(title: str, counts: List[tuple[str, int]]) -> str:
    lines = [f"**{title}**", "", "| Value | Count |", "|---|---:|"]
    for k, v in counts:
        lines.append(f"| {k} | {v} |")
    return "\n".join(lines)


def _read_xlsx_rows_openpyxl(xlsx: Path, sheet: str) -> Tuple[List[str], List[List[object]]]:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet!r} (available: {', '.join(wb.sheetnames)})")
    ws = wb[sheet]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return [], []
    cols = [str(c).strip() if c is not None else "" for c in header]
    rows = [list(r) for r in rows_iter]
    return cols, rows


def _counts_simple(values: Iterable[object], max_rows: int) -> List[Tuple[str, int]]:
    counts: dict[str, int] = {}
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s or s.lower() == "nan":
            continue
        counts[s] = counts.get(s, 0) + 1
    ranked = sorted(counts.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
    return ranked[:max_rows]


def _counts_multi(values: Iterable[object], max_rows: int) -> List[Tuple[str, int]]:
    counts: dict[str, int] = {}
    for v in values:
        if v is None:
            continue
        for item in _split_multi(str(v)):
            if item.lower() == "nan":
                continue
            counts[item] = counts.get(item, 0) + 1
    ranked = sorted(counts.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
    return ranked[:max_rows]


def _phase_rank(phase: str) -> int:
    p = (phase or "").strip().lower()
    # Higher rank => later stage.
    if "launched" in p or "marketed" in p:
        return 100
    if "registered" in p or "pre-registration" in p or "preregistration" in p:
        return 90
    if "phase i/ii" in p or "phase 1/2" in p:
        return 55
    if "phase 4" in p:
        return 80
    if "phase 3" in p:
        return 70
    if "phase 2" in p:
        return 60
    if "phase 1" in p:
        return 50
    if "preclinical" in p or "pre-clinical" in p:
        return 40
    if "discontinued" in p:
        return 20
    if "no development" in p:
        return 10
    if "unknown" in p:
        return 0
    return 30


def _sorted_phases(phases: Iterable[str]) -> List[str]:
    uniq = sorted({(p or "").strip() for p in phases if (p or "").strip() and (p or "").strip().lower() != "nan"})
    # Desc by stage, then alpha for stability.
    uniq.sort(key=lambda x: (_phase_rank(x), x), reverse=True)
    return uniq


def _contingency_indication_by_phase(
    rows: List[List[object]],
    col_index: Dict[str, int],
    *,
    top_indications: int,
    min_indication_count: int,
) -> Tuple[List[str], Dict[str, Dict[str, int]]]:
    """Return (phase_order, counts[indication][phase]=n) for top indications by total."""
    phase_idx = col_index.get("Highest Phase")
    ind_idx = col_index.get("Primary Indications")
    if phase_idx is None or ind_idx is None:
        return [], {}

    totals: Dict[str, int] = {}
    by_ind_phase: Dict[str, Dict[str, int]] = {}
    all_phases: List[str] = []

    for r in rows:
        if phase_idx >= len(r) or ind_idx >= len(r):
            continue
        phase_v = r[phase_idx]
        phase = str(phase_v).strip() if phase_v is not None else ""
        if not phase:
            phase = "Unknown"
        if phase.lower() == "nan":
            phase = "Unknown"
        all_phases.append(phase)

        inds_raw = r[ind_idx]
        if inds_raw is None:
            continue
        inds = _split_multi(str(inds_raw))
        if not inds:
            continue
        for ind in inds:
            if ind.lower() == "nan":
                continue
            totals[ind] = totals.get(ind, 0) + 1
            by_ind_phase.setdefault(ind, {})
            by_ind_phase[ind][phase] = by_ind_phase[ind].get(phase, 0) + 1

    ranked_inds = sorted(totals.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
    top = [k for k, v in ranked_inds if v >= min_indication_count][:top_indications]

    phase_order = _sorted_phases(all_phases)
    filtered: Dict[str, Dict[str, int]] = {}
    for ind in top:
        filtered[ind] = by_ind_phase.get(ind, {})
    return phase_order, filtered


def _to_md_matrix(title: str, phase_order: List[str], by_ind_phase: Dict[str, Dict[str, int]]) -> str:
    if not phase_order or not by_ind_phase:
        return f"**{title}**\n\n_No data available._"
    inds = list(by_ind_phase.keys())
    # header
    lines = [f"**{title}**", "", "| Indication | " + " | ".join(phase_order) + " | Total |", "|---|"
             + "|".join(["---:"] * (len(phase_order) + 1)) + "|"]
    for ind in inds:
        row = []
        total = 0
        for ph in phase_order:
            v = int(by_ind_phase.get(ind, {}).get(ph, 0))
            total += v
            row.append(str(v))
        lines.append("| " + ind + " | " + " | ".join(row) + f" | {total} |")
    return "\n".join(lines)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to gene_cortellis_data.xlsx")
    ap.add_argument("--gene", required=True, help="Gene symbol to filter (e.g., TYK2)")
    ap.add_argument("--sheet", default="Drugs_Comprehensive", help="Sheet name (default: Drugs_Comprehensive)")
    ap.add_argument("--max-rows", type=int, default=20, help="Max rows per table (default: 20)")
    ap.add_argument("--top-indications", type=int, default=12, help="Top N indications for stacked output (default: 12)")
    ap.add_argument("--min-indication-count", type=int, default=2, help="Min count to include an indication (default: 2)")
    ap.add_argument(
        "--format",
        choices=["markdown", "json"],
        default="markdown",
        help="Output format (default: markdown). JSON is intended for report generators.",
    )
    args = ap.parse_args()

    xlsx = Path(args.xlsx).resolve()
    if not xlsx.exists():
        raise SystemExit(f"Not found: {xlsx}")

    try:
        cols, rows = _read_xlsx_rows_openpyxl(xlsx, args.sheet)
    except Exception:
        # Fallback to pandas if openpyxl isn't available in this environment.
        try:
            import pandas as pd  # type: ignore
        except Exception as e:
            raise SystemExit(
                "Missing dependency: install `openpyxl` or `pandas` to read .xlsx in this helper script."
            ) from e

        df = pd.read_excel(xlsx, sheet_name=args.sheet)  # type: ignore[arg-type]
        cols = [str(c) for c in df.columns]
        rows = df.values.tolist()

    col_index = {str(c).strip(): i for i, c in enumerate(cols) if str(c).strip()}
    gene_col = _find_gene_column(cols)
    gene_idx = col_index.get(gene_col) if gene_col else None

    filtered_rows: List[List[object]] = []
    for r in rows:
        rr = list(r)
        if gene_idx is None:
            filtered_rows.append(rr)
            continue
        if gene_idx >= len(rr):
            continue
        gv = rr[gene_idx]
        if str(gv).upper().strip() == args.gene.upper().strip():
            filtered_rows.append(rr)

    def col_values(name: str) -> List[object]:
        idx = col_index.get(name)
        if idx is None:
            return []
        out: List[object] = []
        for r in filtered_rows:
            if idx < len(r):
                out.append(r[idx])
        return out

    highest_phase = _counts_simple(col_values("Highest Phase"), args.max_rows)
    primary_inds = _counts_multi(col_values("Primary Indications"), args.max_rows)
    phase_order, top_ind_phase = _contingency_indication_by_phase(
        filtered_rows,
        col_index,
        top_indications=args.top_indications,
        min_indication_count=args.min_indication_count,
    )

    if args.format == "json":
        payload = {
            "xlsx": str(xlsx),
            "gene": args.gene,
            "sheet": args.sheet,
            "highest_phase_counts": {k: v for k, v in highest_phase},
            "primary_indication_counts": {k: v for k, v in primary_inds},
            "phase_order": phase_order,
            "top_indications_by_phase": top_ind_phase,
            "notes": {
                "primary_indications_split": "split on ';' or '|' (commas are not split)",
                "stacked_bar_intent": "each bar=indication; stacks=highest phase; values are drug-indication counts",
            },
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0

    out_blocks = []
    out_blocks.append(
        _to_md_table("Cortellis: Highest Phase (counts)", highest_phase)
        if highest_phase
        else "_No data for Highest Phase._"
    )
    out_blocks.append("")
    out_blocks.append(
        _to_md_table("Cortellis: Primary Indications (counts)", primary_inds)
        if primary_inds
        else "_No data for Primary Indications._"
    )
    out_blocks.append("")
    out_blocks.append(
        _to_md_matrix(
            "Cortellis: Top indications × Highest Phase (for stacked bar plotting)",
            phase_order,
            top_ind_phase,
        )
    )
    print("\n".join(out_blocks).strip())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
